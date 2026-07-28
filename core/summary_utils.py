"""파일 요약 공통 헬퍼 (포맷·라벨·엑셀 shape)."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

from core.constants import BUDGET_FOOTER_LABELS
from core.pandasai_config import is_total_label

# 하위 호환 alias
FOOTER_LABELS = BUDGET_FOOTER_LABELS

_GRAND_TOTAL_RE = re.compile(
    r"^(?:합\s*계|총\s*계|grand\s*total)$",
    flags=re.IGNORECASE,
)

_SUBTOTAL_RE = re.compile(
    r"^(?:소\s*계|sub\s*total)$",
    flags=re.IGNORECASE,
)


def excel_shape(path: str | Path | None) -> tuple[int, int] | None:
    if not path:
        return None
    excel_path = Path(path)
    if not excel_path.is_file() or excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        shape = (int(worksheet.max_row or 0), int(worksheet.max_column or 0))
        workbook.close()
        if shape[0] > 0 and shape[1] > 0:
            return shape
    except Exception:
        return None
    return None


def is_numeric_col(df: pd.DataFrame, column: object) -> bool:
    return pd.api.types.is_numeric_dtype(df[column])


def cell_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat", "<na>"}:
        return ""
    return text


def compact(text: str) -> str:
    return re.sub(r"\s+", "", text)


def is_grand_total_label(value: object) -> bool:
    text = cell_text(value)
    if not text:
        return False
    return bool(_GRAND_TOTAL_RE.fullmatch(text))


def is_excluded_summary_label(value: object) -> bool:
    text = cell_text(value)
    if not text:
        return False
    if is_total_label(text) or is_grand_total_label(text) or _SUBTOTAL_RE.fullmatch(text):
        return True
    return compact(text) in {compact(label) for label in BUDGET_FOOTER_LABELS}


def fmt_won(value: float) -> str:
    number = int(round(value))
    return f"{number:,}원"


def fmt_number(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}"
    return f"{value:,.2f}"
