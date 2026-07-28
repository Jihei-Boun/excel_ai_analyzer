"""파일 요약 요청용 규칙 기반 분석 (예실대비표 특화 + 범용 폴백)."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

from core.pandasai_config import is_total_label, prepare_dataframe_for_ai

_SUMMARY_KEYWORDS = (
    "요약",
    "개요",
    "파일소개",
    "파일설명",
    "어떤파일",
    "무슨파일",
    "파일내용",
    "파일알려",
    "summarize",
    "summary",
    "overview",
)

_BUDGET_COLUMN_HINTS = (
    "계획예산",
    "실행예산",
    "예산잔액",
    "집행계",
    "당년도집행",
    "가집행",
    "비목분류",
)

_FOOTER_LABELS = (
    "내부흡수액",
    "외부유출액",
)

_GRAND_TOTAL_RE = re.compile(
    r"^(?:합\s*계|총\s*계|grand\s*total)$",
    flags=re.IGNORECASE,
)

_SUBTOTAL_RE = re.compile(
    r"^(?:소\s*계|sub\s*total)$",
    flags=re.IGNORECASE,
)


def is_summary_request(prompt: str) -> bool:
    """파일 요약·개요 요청인지 판별한다. 차트 요청은 제외."""
    if not prompt or not prompt.strip():
        return False
    lowered = prompt.lower()
    if any(k in lowered for k in ("차트", "그래프", "chart", "plot", "graph")):
        return False
    normalized = re.sub(r"\s+", "", lowered)
    return any(keyword in normalized for keyword in _SUMMARY_KEYWORDS)


def build_file_summary(
    df: pd.DataFrame,
    *,
    file_name: str | None = None,
    sheet_name: str | None = None,
    sheet_names: list[str] | None = None,
    file_path: str | Path | None = None,
) -> str:
    """DataFrame을 읽어 사람이 읽을 수 있는 파일 요약 문장을 만든다."""
    if df is None or df.empty:
        return "데이터가 비어 있어 요약할 내용이 없습니다."

    prepared = prepare_dataframe_for_ai(df)
    sheets = sheet_names or ([sheet_name] if sheet_name else [])
    excel_shape = _excel_shape(file_path) if file_path else None

    if _looks_like_budget_table(prepared):
        return _build_budget_summary(
            prepared,
            file_name=file_name,
            sheet_name=sheet_name,
            sheets=sheets,
            excel_shape=excel_shape,
        )
    return _build_generic_summary(
        prepared,
        file_name=file_name,
        sheet_name=sheet_name,
        sheets=sheets,
        excel_shape=excel_shape,
    )


def build_multi_file_summary(
    named_dfs: list[tuple[str, pd.DataFrame]],
    *,
    sheet_info: dict[str, dict] | None = None,
) -> str:
    """여러 파일을 짧게 이어서 요약한다."""
    if not named_dfs:
        return "요약할 파일이 없습니다."

    parts: list[str] = [f"선택된 파일 {len(named_dfs)}개를 요약합니다.\n"]
    for name, frame in named_dfs:
        info = (sheet_info or {}).get(name) or {}
        block = build_file_summary(
            frame,
            file_name=name,
            sheet_name=info.get("current_sheet"),
            sheet_names=info.get("sheet_names"),
            file_path=info.get("path"),
        )
        parts.append(f"### {name}\n{block}")
    return "\n\n".join(parts)


def _build_budget_summary(
    df: pd.DataFrame,
    *,
    file_name: str | None,
    sheet_name: str | None,
    sheets: list[str],
    excel_shape: tuple[int, int] | None,
) -> str:
    _ = file_name
    rows, cols = excel_shape or (len(df), len(df.columns))
    sheet_label = sheet_name or (sheets[0] if sheets else "Sheet1")
    sheet_count = len(sheets) if sheets else 1

    item_col = _pick_column(df, ("비용명_2", "비용명", "항목", "세목", "내역"))
    category_col = _pick_column(df, ("비목분류", "대분류", "중분류", "분류", "구분"))
    budget_col = _pick_column(
        df,
        ("실행예산_합계", "계획예산", "당년도예산", "실행예산"),
    )
    executed_col = _pick_column(
        df,
        ("집행계_합계", "당해누계", "당년도집행", "집행계"),
    )
    remaining_col = _pick_column(df, ("예산잔액_합계", "예산잔액"))
    current_remaining_col = _pick_column(
        df,
        ("예산잔액_당해잔액", "당해잔액"),
    )

    total_row = _find_grand_total_row(df, category_col)
    detail = _detail_rows(df, category_col)

    budget = _row_or_sum(total_row, detail, budget_col)
    executed = _row_or_sum(total_row, detail, executed_col)
    remaining = _row_or_sum(total_row, detail, remaining_col)
    rate = (executed / budget * 100.0) if budget and budget > 0 and executed is not None else None

    categories = _major_categories(df, category_col)
    top_executed = _top_items(detail, item_col, executed_col, n=3)
    top_remaining = _top_items(detail, item_col, remaining_col, n=3)
    negatives = _negative_items(detail, item_col, current_remaining_col or remaining_col)
    key_columns = _key_budget_columns(df.columns)

    lines: list[str] = []
    lines.append(
        "이 파일은 연구과제의 예산 계획, 집행 내역, 가집행 금액 및 잔액을 "
        "관리하는 예실대비표입니다."
    )
    lines.append("")
    lines.append(f"* 시트 수: {sheet_count}개 (`{sheet_label}`)")
    lines.append(f"* 실제 데이터 범위: {rows}행 × {cols}열")
    if categories:
        lines.append(f"* 주요 항목: {', '.join(categories)}")
    if budget is not None:
        lines.append(f"* 전체 예산: **{_fmt_won(budget)}**")
    if executed is not None:
        lines.append(f"* 누적 집행액: **{_fmt_won(executed)}**")
    if remaining is not None:
        lines.append(f"* 전체 예산잔액: **{_fmt_won(remaining)}**")
    if rate is not None:
        lines.append(f"* 전체 집행률: 약 **{rate:.1f}%**")

    if top_executed:
        lines.append("")
        lines.append(_format_ranking_sentence("집행액은", top_executed, "으로 가장 크고", "순입니다."))

    if top_remaining:
        lines.append("")
        lines.append(_format_ranking_sentence("잔액이 큰 항목은", top_remaining, "", "입니다."))

    if negatives:
        lines.append("")
        names = "와 ".join(name for name, _ in negatives) if len(negatives) == 2 else ", ".join(
            name for name, _ in negatives
        )
        amounts = ", ".join(f"**{_fmt_won(value)}**" for _, value in negatives)
        if len(negatives) == 2:
            amounts = f"각각 {amounts}"
        balance_label = "당해 잔액" if current_remaining_col else "예산잔액"
        lines.append(
            f"또한 {names}의 {balance_label}이 {amounts}으로 음수이므로 "
            "예산 초과 또는 이월·가집행 처리 여부를 확인할 필요가 있습니다."
        )

    if key_columns:
        lines.append("")
        col_list = ", ".join(f"`{name}`" for name in key_columns)
        footer_note = ""
        if _has_footer_labels(df, category_col):
            footer_note = (
                " 하단에는 내부흡수액, 외부유출액, 전체 합계가 정리되어 있습니다."
            )
        lines.append(
            f"표에는 {col_list} 등의 컬럼이 있으며,{footer_note}"
            if footer_note
            else f"표에는 {col_list} 등의 컬럼이 있습니다."
        )

    return "\n".join(lines).rstrip()


def _build_generic_summary(
    df: pd.DataFrame,
    *,
    file_name: str | None,
    sheet_name: str | None,
    sheets: list[str],
    excel_shape: tuple[int, int] | None,
) -> str:
    rows, cols = excel_shape or (len(df), len(df.columns))
    sheet_label = sheet_name or (sheets[0] if sheets else None)
    sheet_count = len(sheets) if sheets else 1
    numeric_cols = [
        col
        for col in df.columns
        if pd.to_numeric(df[col], errors="coerce").notna().any()
    ]
    text_cols = [col for col in df.columns if col not in numeric_cols]

    lines = [
        (
            f"이 파일(`{file_name}`)은 {rows}행 × {cols}열 표 형태 데이터입니다."
            if file_name
            else f"이 파일은 {rows}행 × {cols}열 표 형태 데이터입니다."
        ),
        "",
        f"* 시트 수: {sheet_count}개"
        + (f" (`{sheet_label}`)" if sheet_label else ""),
        f"* 실제 데이터 범위: {rows}행 × {cols}열",
        f"* 수치형 컬럼: {len(numeric_cols)}개",
        f"* 문자형 컬럼: {len(text_cols)}개",
    ]
    preview_cols = [str(c) for c in list(df.columns)[:8]]
    more = "" if len(df.columns) <= 8 else f" 외 {len(df.columns) - 8}개"
    lines.append(f"* 주요 컬럼: {', '.join(f'`{c}`' for c in preview_cols)}{more}")

    if numeric_cols:
        lines.append("")
        lines.append("주요 수치 컬럼 합계(단순 합산):")
        for col in numeric_cols[:5]:
            total = float(pd.to_numeric(df[col], errors="coerce").sum(skipna=True))
            lines.append(f"* `{col}`: {_fmt_number(total)}")

    return "\n".join(lines)


def _looks_like_budget_table(df: pd.DataFrame) -> bool:
    joined = " ".join(str(c) for c in df.columns)
    hits = sum(1 for hint in _BUDGET_COLUMN_HINTS if hint in joined)
    return hits >= 2


def _pick_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    columns = [str(c) for c in df.columns]
    for wanted in candidates:
        for column in columns:
            if column == wanted:
                return column
    for wanted in candidates:
        for column in columns:
            if wanted in column:
                return column
    return None


def _find_grand_total_row(
    df: pd.DataFrame,
    category_col: str | None,
) -> pd.Series | None:
    """하단 최종 합계 행을 찾는다. 소계·내부흡수·외부유출은 제외."""
    for index in reversed(range(len(df))):
        row = df.iloc[index]
        labels = [_cell_text(row.get(col)) for col in df.columns if not _is_numeric_col(df, col)]
        if category_col:
            labels.insert(0, _cell_text(row.get(category_col)))
        if any(_is_grand_total_label(label) for label in labels if label):
            return row
    return None


def _detail_rows(df: pd.DataFrame, category_col: str | None) -> pd.DataFrame:
    """소계·합계·내부흡수액·외부유출액 행을 제외한 세부 항목."""
    mask = pd.Series(True, index=df.index)
    for column in df.columns:
        if _is_numeric_col(df, column):
            continue
        mask &= ~df[column].map(_is_excluded_summary_label)
    if category_col and category_col in df.columns:
        mask &= ~df[category_col].map(_is_excluded_summary_label)
    return df.loc[mask].copy()


def _major_categories(df: pd.DataFrame, category_col: str | None) -> list[str]:
    if not category_col or category_col not in df.columns:
        return []
    values: list[str] = []
    seen: set[str] = set()
    for value in df[category_col].tolist():
        text = _cell_text(value)
        if not text or _is_excluded_summary_label(text):
            continue
        if text in seen:
            continue
        seen.add(text)
        values.append(text)
    return values


def _top_items(
    detail: pd.DataFrame,
    item_col: str | None,
    metric_col: str | None,
    *,
    n: int,
) -> list[tuple[str, float]]:
    if detail.empty or not item_col or not metric_col:
        return []
    if item_col not in detail.columns or metric_col not in detail.columns:
        return []

    work = detail[[item_col, metric_col]].copy()
    work[metric_col] = pd.to_numeric(work[metric_col], errors="coerce")
    work[item_col] = work[item_col].map(_cell_text)
    work = work.dropna(subset=[metric_col])
    work = work[work[item_col].astype(bool)]
    work = work[~work[item_col].map(_is_excluded_summary_label)]
    if work.empty:
        return []

    grouped = (
        work.groupby(item_col, as_index=False)[metric_col]
        .sum()
        .sort_values(metric_col, ascending=False)
        .head(n)
    )
    return [
        (str(row[item_col]), float(row[metric_col]))
        for _, row in grouped.iterrows()
    ]


def _negative_items(
    detail: pd.DataFrame,
    item_col: str | None,
    metric_col: str | None,
) -> list[tuple[str, float]]:
    if detail.empty or not item_col or not metric_col:
        return []
    if item_col not in detail.columns or metric_col not in detail.columns:
        return []

    items: list[tuple[str, float]] = []
    for _, row in detail.iterrows():
        name = _cell_text(row.get(item_col))
        if not name or _is_excluded_summary_label(name):
            continue
        value = pd.to_numeric(row.get(metric_col), errors="coerce")
        if pd.isna(value) or float(value) >= 0:
            continue
        items.append((name, float(value)))
    return items


def _row_or_sum(
    total_row: pd.Series | None,
    detail: pd.DataFrame,
    column: str | None,
) -> float | None:
    if not column:
        return None
    if total_row is not None and column in total_row.index:
        value = pd.to_numeric(total_row.get(column), errors="coerce")
        if pd.notna(value):
            return float(value)
    if column not in detail.columns:
        return None
    total = pd.to_numeric(detail[column], errors="coerce").sum(skipna=True)
    if pd.isna(total):
        return None
    return float(total)


def _key_budget_columns(columns: pd.Index) -> list[str]:
    preferred = (
        "계획예산",
        "실행예산",
        "당년도집행",
        "가집행금액",
        "집행계",
        "예산잔액",
    )
    found: list[str] = []
    for hint in preferred:
        for column in columns:
            text = str(column)
            base = text.split("_", 1)[0]
            if hint in (text, base) or text.startswith(f"{hint}_"):
                if hint not in found:
                    found.append(hint)
                break
    return found


def _has_footer_labels(df: pd.DataFrame, category_col: str | None) -> bool:
    if not category_col or category_col not in df.columns:
        return False
    labels = {_compact(_cell_text(v)) for v in df[category_col].tolist()}
    return any(_compact(label) in labels for label in _FOOTER_LABELS)


def _format_ranking_sentence(
    prefix: str,
    items: list[tuple[str, float]],
    first_suffix: str,
    end_suffix: str,
) -> str:
    if not items:
        return ""
    first_name, first_value = items[0]
    if len(items) == 1:
        return f"{prefix} {first_name}가 {_fmt_won(first_value)}{first_suffix or end_suffix}."
    rest = ", ".join(f"{name} {_fmt_won(value)}" for name, value in items[1:])
    if first_suffix:
        return f"{prefix} {first_name}가 {_fmt_won(first_value)}{first_suffix}, {rest} {end_suffix}"
    return f"{prefix} {first_name} {_fmt_won(first_value)}, {rest}{end_suffix}"


def _excel_shape(path: str | Path | None) -> tuple[int, int] | None:
    if not path:
        return None
    excel_path = Path(path)
    if not excel_path.is_file() or excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        shape = (int(worksheet.max_row or 0), int(worksheet.max_column or 0))
        workbook.close()
        if shape[0] > 0 and shape[1] > 0:
            return shape
    except Exception:
        return None
    return None


def _is_grand_total_label(value: object) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    return bool(_GRAND_TOTAL_RE.fullmatch(text))


def _is_excluded_summary_label(value: object) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    if is_total_label(text) or _is_grand_total_label(text) or _SUBTOTAL_RE.fullmatch(text):
        return True
    compact = _compact(text)
    return compact in {_compact(label) for label in _FOOTER_LABELS}


def _is_numeric_col(df: pd.DataFrame, column: object) -> bool:
    return pd.api.types.is_numeric_dtype(df[column])


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat", "<na>"}:
        return ""
    return text


def _compact(text: str) -> str:
    return re.sub(r"\s+", "", text)


def _fmt_won(value: float) -> str:
    number = int(round(value))
    return f"{number:,}원"


def _fmt_number(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}"
    return f"{value:,.2f}"
