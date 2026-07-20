"""엑셀 파일을 범용 pandas DataFrame으로 읽고 정리한다."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

_UNNAMED_HEADER_RE = re.compile(r"^Unnamed:\s*\d+$", re.IGNORECASE)
_MERGED_SUFFIX_RE = re.compile(r"^(.+)_(\d+)$")


def load_excel(path: str | Path, *, sheet_name: str | int = 0) -> pd.DataFrame:
    """엑셀 파일을 읽고 기본 전처리를 적용한다."""
    df = pd.read_excel(path, sheet_name=sheet_name)
    df = _apply_merged_header_names(path, sheet_name, df)
    return sanitize_dataframe(df)


def find_merged_header_pair(
    columns: pd.Index | list[str],
    base_name: str | None,
) -> tuple[str, str] | None:
    """병합 헤더로 생긴 인접 컬럼 쌍 (왼쪽=코드, 오른쪽=명칭)을 찾는다."""
    cols = [str(column) for column in columns]
    pairs: list[tuple[str, str]] = []
    for index, column in enumerate(cols[:-1]):
        suffix = cols[index + 1]
        match = _MERGED_SUFFIX_RE.fullmatch(suffix)
        if match and match.group(1) == column:
            pairs.append((column, suffix))

    if not pairs:
        return None
    if not base_name:
        return pairs[0]

    normalized_base = merged_header_base(base_name)
    for left, right in pairs:
        if base_name in (left, right) or normalized_base == left:
            return (left, right)
    return None


def merged_header_base(name: str) -> str:
    """'비용명_2'처럼 병합 헤더 접미사가 붙은 컬럼명의 기본 이름을 반환한다."""
    match = _MERGED_SUFFIX_RE.fullmatch(str(name))
    return match.group(1) if match else str(name)


def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """원본 의미를 유지하며 컬럼명·빈 행·안전한 숫자 문자열만 정리한다."""
    return _preprocess(df)


def _preprocess(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = _propagate_unnamed_headers(df.columns)
    df.columns = _unique_column_names(df.columns)
    df = df.dropna(how="all").reset_index(drop=True)

    for col in df.columns:
        df[col] = _coerce_column(df[col])

    return df


def _apply_merged_header_names(
    path: str | Path,
    sheet_name: str | int,
    df: pd.DataFrame,
) -> pd.DataFrame:
    """엑셀 1행 병합 셀 헤더를 모든 열에 전파한다."""
    excel_path = Path(path)
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return df

    try:
        from openpyxl import load_workbook
    except ImportError:
        return df

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        if isinstance(sheet_name, int):
            worksheet = workbook.worksheets[sheet_name]
        else:
            worksheet = workbook[sheet_name]

        names = [str(column).strip() for column in df.columns]
        for merge_range in worksheet.merged_cells.ranges:
            if merge_range.min_row != 1 or merge_range.max_row != 1:
                continue
            if merge_range.min_col == merge_range.max_col:
                continue

            header_value = worksheet.cell(merge_range.min_row, merge_range.min_col).value
            header_name = str(header_value).strip() if header_value is not None else ""
            if not header_name:
                continue

            for col_index in range(merge_range.min_col - 1, merge_range.max_col):
                if 0 <= col_index < len(names):
                    names[col_index] = header_name

        workbook.close()
        df = df.copy()
        df.columns = names
    except Exception:
        return df

    return df


def _propagate_unnamed_headers(columns: pd.Index) -> list[str]:
    """'Unnamed' 열이 바로 앞 열과 병합 헤더였을 때 같은 이름을 부여한다."""
    names: list[str] = []
    for index, column in enumerate(columns, start=1):
        text = str(column).strip()
        if _is_unnamed_header(text) and names:
            previous = names[-1]
            if previous and not _is_unnamed_header(previous):
                names.append(previous)
                continue
        names.append(text or f"column_{index}")
    return names


def _is_unnamed_header(name: str) -> bool:
    if not name:
        return True
    if name.startswith("column_"):
        return True
    return bool(_UNNAMED_HEADER_RE.fullmatch(name))


def _unique_column_names(columns: pd.Index) -> list[str]:
    names: list[str] = []
    counts: dict[str, int] = {}
    for index, column in enumerate(columns, start=1):
        base = str(column).strip() or f"column_{index}"
        count = counts.get(base, 0)
        counts[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def _coerce_column(series: pd.Series) -> pd.Series:
    """혼합 타입·콤마 숫자를 정리해 Arrow/연산에 맞게 만든다."""
    if pd.api.types.is_bool_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return series

    if pd.api.types.is_numeric_dtype(series):
        return series

    numeric = _to_numeric_series(series)
    if numeric is not None:
        return numeric

    return _to_clean_string_series(series)


def _to_numeric_series(series: pd.Series) -> pd.Series | None:
    cleaned = (
        series.map(_cell_to_text)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA})
    numeric = pd.to_numeric(cleaned, errors="coerce")

    non_empty = cleaned.notna().sum()
    if non_empty == 0:
        return None

    converted = numeric.notna().sum()
    if converted == non_empty:
        return numeric
    return None


def _to_clean_string_series(series: pd.Series) -> pd.Series:
    return series.map(_clean_string_value).astype("string")


def _clean_string_value(value: object) -> object:
    if value is None:
        return pd.NA
    try:
        if pd.isna(value):
            return pd.NA
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
