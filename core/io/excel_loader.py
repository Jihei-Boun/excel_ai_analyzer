"""엑셀 파일을 범용 pandas DataFrame으로 읽고 정리한다."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd


_UNNAMED_HEADER_RE = re.compile(r"^Unnamed(\s*:\s*\d+)?(_level_\d+)?$", re.IGNORECASE)
_UNNAMED_TOKEN_RE = re.compile(r"^Unnamed", re.IGNORECASE)
_MERGED_SUFFIX_RE = re.compile(r"^(.+)_(\d+)$")
_COMPOUND_HEADER_RE = re.compile(r"^(.+)_(.+)$")
# 본문 날짜/일시 문자열 (헤더 하위분류와 구분)
_DATE_VALUE_RE = re.compile(
    r"^("
    r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}"
    r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?)?"
    r"|"
    r"\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}"
    r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?)?"
    r")$"
)


def load_excel(path: str | Path, *, sheet_name: str | int = 0) -> pd.DataFrame:
    """엑셀 파일을 읽고 기본 전처리를 적용한다."""
    excel_path = Path(path)
    header_depth = _detect_header_depth(excel_path, sheet_name)

    if header_depth >= 2:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=[0, 1])
        df = _flatten_two_level_headers(excel_path, sheet_name, df)
    else:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df = _apply_merged_header_names(excel_path, sheet_name, df)

    return sanitize_dataframe(df)


def load_csv(path: str | Path, *, encoding: str | None = None) -> pd.DataFrame:
    """CSV를 읽고 엑셀과 동일한 sanitize를 적용한다."""
    csv_path = Path(path)
    tried: list[str] = []
    encodings = [encoding] if encoding else ["utf-8-sig", "utf-8", "cp949", "euc-kr"]
    last_error: Exception | None = None
    for enc in encodings:
        if not enc or enc in tried:
            continue
        tried.append(enc)
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            return sanitize_dataframe(df)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise ValueError(f"CSV를 읽을 수 없습니다: {csv_path.name}")


def is_csv_path(path: str | Path) -> bool:
    return Path(path).suffix.lower() == ".csv"


def is_excel_path(path: str | Path) -> bool:
    return Path(path).suffix.lower() in {".xlsx", ".xls", ".xlsm"}


# CSV는 시트가 없으므로 UI·캐시용 가상 시트명
CSV_SHEET_NAME = "CSV"


def load_tabular(
    path: str | Path,
    *,
    sheet_name: str | int | None = 0,
) -> pd.DataFrame:
    """엑셀 또는 CSV를 DataFrame으로 로드한다."""
    file_path = Path(path)
    if is_csv_path(file_path):
        return load_csv(file_path)
    if sheet_name is None:
        sheet_name = 0
    return load_excel(file_path, sheet_name=sheet_name)


def find_merged_header_pair(
    columns: pd.Index | list[str],
    base_name: str | None,
) -> tuple[str, str] | None:
    """병합 헤더로 생긴 인접 컬럼 쌍 (왼쪽=코드, 오른쪽=명칭)을 찾는다."""
    cols = [str(column) for column in columns]
    pairs: list[tuple[str, str]] = []
    for index, column in enumerate(cols[:-1]):
        suffix = cols[index + 1]
        match = _MERGED_SUFFIX_RE.fullmatch(suffix)
        if match and match.group(1) == column:
            pairs.append((column, suffix))

    if not pairs:
        return None
    if not base_name:
        return pairs[0]

    normalized_base = merged_header_base(base_name)
    for left, right in pairs:
        if base_name in (left, right) or normalized_base == left:
            return (left, right)
    return None


def merged_header_base(name: str) -> str:
    """병합/복합 헤더 컬럼명의 상위 이름을 반환한다.

    예: '실행예산_2' → '실행예산', '실행예산_이월예산' → '실행예산'
    """
    text = str(name)
    match = _MERGED_SUFFIX_RE.fullmatch(text)
    if match:
        return match.group(1)
    if "_" in text:
        return text.split("_", 1)[0]
    return text


def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """원본 의미를 유지하며 컬럼명·빈 행·안전한 숫자 문자열만 정리한다.

    이후 공통 정규화(`normalize_dataframe`)를 적용해 업로드 편차를 흡수한다.
    """
    from core.io.normalize import normalize_dataframe

    return normalize_dataframe(_preprocess(df))


def _preprocess(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = _flatten_multiindex_names(df.columns)
    df.columns = _propagate_unnamed_headers(df.columns)
    df.columns = _unique_column_names(df.columns)
    df = df.dropna(how="all").reset_index(drop=True)

    for col in df.columns:
        df[col] = _coerce_column(df[col])

    return df


def _detect_header_depth(path: Path, sheet_name: str | int) -> int:
    """2단 헤더(상위 병합 + 하위 세부분류)면 2, 아니면 1."""
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        return 1

    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
        worksheet = (
            workbook.worksheets[sheet_name]
            if isinstance(sheet_name, int)
            else workbook[sheet_name]
        )
        max_col = worksheet.max_column or 0
        if max_col < 1:
            workbook.close()
            return 1

        row1 = [_cell_text(worksheet.cell(1, col).value) for col in range(1, max_col + 1)]
        row2 = [_cell_text(worksheet.cell(2, col).value) for col in range(1, max_col + 1)]

        # 1행 가로 병합(상위 카테고리)이 있어야 2단 헤더 후보
        has_parent_merge = False
        for merge_range in worksheet.merged_cells.ranges:
            if (
                merge_range.min_row == 1
                and merge_range.max_row == 1
                and merge_range.max_col > merge_range.min_col
            ):
                has_parent_merge = True
                value = _cell_text(
                    worksheet.cell(merge_range.min_row, merge_range.min_col).value
                )
                if not value:
                    continue
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    index = col - 1
                    if 0 <= index < len(row1):
                        row1[index] = value

        workbook.close()
    except Exception:
        return 1

    if not has_parent_merge:
        return 1

    meaningful_children = sum(
        1
        for parent, child in zip(row1, row2)
        if child
        and not _is_unnamed_header(child)
        and not _looks_like_data_value(child)
        and (not parent or parent != child)
    )
    # 상위 병합 + 하위에 서로 다른 세부분류가 2개 이상이면 2단 헤더
    if meaningful_children >= 2:
        return 2
    return 1


def _flatten_two_level_headers(
    path: Path,
    sheet_name: str | int,
    df: pd.DataFrame,
) -> pd.DataFrame:
    """openpyxl로 1·2행을 읽어 parent_child 형태의 컬럼명을 만든다."""
    parents, children = _read_header_rows(path, sheet_name, expected_cols=len(df.columns))
    if not parents:
        df = df.copy()
        df.columns = _flatten_multiindex_names(df.columns)
        return df

    # pandas MultiIndex 길이보다 openpyxl 열이 길/짧을 수 있어 맞춤
    width = len(df.columns)
    parents = _pad_or_trim(parents, width)
    children = _pad_or_trim(children, width)
    parents = _forward_fill_names(parents)

    names: list[str] = []
    for parent, child in zip(parents, children):
        names.append(_compose_header_name(parent, child))

    result = df.copy()
    result.columns = names
    return result


def _read_header_rows(
    path: Path,
    sheet_name: str | int,
    *,
    expected_cols: int,
) -> tuple[list[str], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], []

    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
        worksheet = (
            workbook.worksheets[sheet_name]
            if isinstance(sheet_name, int)
            else workbook[sheet_name]
        )
        max_col = max(expected_cols, worksheet.max_column or expected_cols)
        parents = [_cell_text(worksheet.cell(1, col).value) for col in range(1, max_col + 1)]
        children = [_cell_text(worksheet.cell(2, col).value) for col in range(1, max_col + 1)]

        # 1행 병합 셀 값을 병합 구간에 전파
        for merge_range in worksheet.merged_cells.ranges:
            if merge_range.min_row != 1 or merge_range.max_row != 1:
                continue
            value = _cell_text(worksheet.cell(merge_range.min_row, merge_range.min_col).value)
            if not value:
                continue
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                index = col - 1
                if 0 <= index < len(parents):
                    parents[index] = value

        workbook.close()
        return parents[:expected_cols], children[:expected_cols]
    except Exception:
        return [], []


def _compose_header_name(parent: str, child: str) -> str:
    parent = parent.strip()
    child = child.strip()
    parent_ok = bool(parent) and not _is_unnamed_header(parent)
    child_ok = (
        bool(child)
        and not _is_unnamed_header(child)
        and not _looks_like_data_value(child)
    )

    if parent_ok and child_ok:
        if parent == child:
            return parent
        # 이미 parent_child 형태면 그대로
        if child.startswith(f"{parent}_"):
            return child
        return f"{parent}_{child}"
    if parent_ok:
        return parent
    if child_ok:
        return child
    return parent or child or "column"


def _flatten_multiindex_names(columns: pd.Index) -> list[str]:
    if not isinstance(columns, pd.MultiIndex):
        return [str(column) for column in columns]

    level0 = [_clean_header_token(value) for value in columns.get_level_values(0)]
    level1 = [_clean_header_token(value) for value in columns.get_level_values(1)]
    level0 = _forward_fill_names(level0)
    return [_compose_header_name(parent, child) for parent, child in zip(level0, level1)]


def _clean_header_token(value: object) -> str:
    text = _cell_text(value)
    if _is_unnamed_header(text) or _UNNAMED_TOKEN_RE.match(text):
        return ""
    return text


def _forward_fill_names(names: list[str]) -> list[str]:
    filled: list[str] = []
    last = ""
    for name in names:
        text = name.strip()
        if text and not _is_unnamed_header(text):
            last = text
            filled.append(text)
        else:
            filled.append(last)
    return filled


def _pad_or_trim(values: list[str], width: int) -> list[str]:
    if len(values) >= width:
        return values[:width]
    return values + [""] * (width - len(values))


def _looks_like_data_value(text: str) -> bool:
    """헤더가 아니라 본문 값처럼 보이면 True."""
    stripped = text.strip()
    if not stripped:
        return False
    if _DATE_VALUE_RE.fullmatch(stripped):
        return True
    cleaned = stripped.replace(",", "").replace(" ", "")
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _apply_merged_header_names(
    path: str | Path,
    sheet_name: str | int,
    df: pd.DataFrame,
) -> pd.DataFrame:
    """엑셀 1행 병합 셀 헤더를 모든 열에 전파한다."""
    excel_path = Path(path)
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return df

    try:
        from openpyxl import load_workbook
    except ImportError:
        return df

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        if isinstance(sheet_name, int):
            worksheet = workbook.worksheets[sheet_name]
        else:
            worksheet = workbook[sheet_name]

        names = [str(column).strip() for column in df.columns]
        for merge_range in worksheet.merged_cells.ranges:
            if merge_range.min_row != 1 or merge_range.max_row != 1:
                continue
            if merge_range.min_col == merge_range.max_col:
                continue

            header_value = worksheet.cell(merge_range.min_row, merge_range.min_col).value
            header_name = str(header_value).strip() if header_value is not None else ""
            if not header_name:
                continue

            for col_index in range(merge_range.min_col - 1, merge_range.max_col):
                if 0 <= col_index < len(names):
                    names[col_index] = header_name

        workbook.close()
        df = df.copy()
        df.columns = names
    except Exception:
        return df

    return df


def _propagate_unnamed_headers(columns: pd.Index) -> list[str]:
    """'Unnamed' 열이 바로 앞 열과 병합 헤더였을 때 같은 이름을 부여한다."""
    names: list[str] = []
    for index, column in enumerate(columns, start=1):
        text = str(column).strip()
        if _is_unnamed_header(text) and names:
            previous = names[-1]
            if previous and not _is_unnamed_header(previous):
                names.append(previous)
                continue
        names.append(text or f"column_{index}")
    return names


def _is_unnamed_header(name: str) -> bool:
    if not name:
        return True
    if name.startswith("column_"):
        return True
    if _UNNAMED_HEADER_RE.fullmatch(name):
        return True
    return bool(_UNNAMED_TOKEN_RE.match(name))


def _unique_column_names(columns: pd.Index) -> list[str]:
    names: list[str] = []
    counts: dict[str, int] = {}
    for index, column in enumerate(columns, start=1):
        base = str(column).strip() or f"column_{index}"
        count = counts.get(base, 0)
        counts[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def _coerce_column(series: pd.Series) -> pd.Series:
    """혼합 타입·콤마 숫자를 정리해 Arrow/연산에 맞게 만든다."""
    if pd.api.types.is_bool_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return series

    if pd.api.types.is_numeric_dtype(series):
        return series

    numeric = _to_numeric_series(series)
    if numeric is not None:
        return numeric

    return _to_clean_string_series(series)


def _to_numeric_series(series: pd.Series) -> pd.Series | None:
    cleaned = (
        series.map(_cell_to_text)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA})
    numeric = pd.to_numeric(cleaned, errors="coerce")

    non_empty = cleaned.notna().sum()
    if non_empty == 0:
        return None

    converted = numeric.notna().sum()
    if converted == non_empty:
        return numeric
    return None


def _to_clean_string_series(series: pd.Series) -> pd.Series:
    return series.map(_clean_string_value).astype("string")


def _clean_string_value(value: object) -> object:
    if value is None:
        return pd.NA
    try:
        if pd.isna(value):
            return pd.NA
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_text(value: object) -> str:
    return _cell_to_text(value)
