"""엑셀 전처리 단위 테스트."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from core.excel_loader import load_excel, merged_header_base, sanitize_dataframe


def test_sanitize_drops_all_empty_rows() -> None:
    cleaned = sanitize_dataframe(
        pd.DataFrame(
            {
                "항목": ["A", None, "B"],
                "금액": [1, None, 2],
                "비고": [None, None, None],
            }
        )
    )
    assert list(cleaned.columns) == ["항목", "금액", "비고"]
    assert len(cleaned) == 2
    assert cleaned.iloc[0]["항목"] == "A"


def test_sanitize_makes_duplicate_headers_unique() -> None:
    df = pd.DataFrame([[1, 2], [3, 4]])
    df.columns = ["코드", "코드"]
    cleaned = sanitize_dataframe(df)
    assert cleaned.columns.tolist() == ["코드", "코드_2"]


def test_load_excel_reads_basic_sheet(tmp_path: Path) -> None:
    path = tmp_path / "sample.xlsx"
    pd.DataFrame({"이름": ["갑", "을"], "값": [10, 20]}).to_excel(path, index=False)
    loaded = load_excel(path)
    assert list(loaded.columns) == ["이름", "값"]
    assert len(loaded) == 2
    assert int(loaded.iloc[0]["값"]) == 10


def test_load_excel_flattens_two_level_merged_headers(tmp_path: Path) -> None:
    """실행예산 아래 이월/당해/합계 → 실행예산_이월예산 형태."""
    path = tmp_path / "budget.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "비목분류"
    sheet["B1"] = "계획예산"
    sheet["C1"] = "실행예산"
    sheet.merge_cells("C1:E1")
    sheet["A2"] = None
    sheet["B2"] = None
    sheet["C2"] = "이월예산"
    sheet["D2"] = "당해예산"
    sheet["E2"] = "합계"
    sheet["A3"] = "인건비"
    sheet["B3"] = 100
    sheet["C3"] = 10
    sheet["D3"] = 20
    sheet["E3"] = 30
    workbook.save(path)

    loaded = load_excel(path)
    assert list(loaded.columns) == [
        "비목분류",
        "계획예산",
        "실행예산_이월예산",
        "실행예산_당해예산",
        "실행예산_합계",
    ]
    assert loaded.iloc[0]["비목분류"] == "인건비"
    assert int(loaded.iloc[0]["실행예산_합계"]) == 30


def test_load_excel_keeps_single_header_when_row2_is_data(tmp_path: Path) -> None:
    """1행만 헤더이고 2행이 텍스트 본문이면 이중 헤더로 오인하지 않는다."""
    from datetime import datetime

    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    headers = ["판매일", "지역", "상품", "담당자", "수량", "단가", "매출", "결제수단"]
    for col, name in enumerate(headers, start=1):
        sheet.cell(1, col, name)
    sheet.cell(2, 1, datetime(2026, 1, 1))
    sheet.cell(2, 2, "서울")
    sheet.cell(2, 3, "A")
    sheet.cell(2, 4, "김철수")
    sheet.cell(2, 5, 2)
    sheet.cell(2, 6, 100000)
    sheet.cell(2, 7, 200000)
    sheet.cell(2, 8, "카드")
    sheet.cell(3, 1, datetime(2026, 1, 2))
    sheet.cell(3, 2, "서울")
    sheet.cell(3, 3, "B")
    sheet.cell(3, 4, "이영희")
    sheet.cell(3, 5, 3)
    sheet.cell(3, 6, 115000)
    sheet.cell(3, 7, 345000)
    sheet.cell(3, 8, "계좌이체")
    workbook.save(path)

    loaded = load_excel(path)
    assert list(loaded.columns) == headers
    assert len(loaded) == 2
    assert loaded.iloc[0]["지역"] == "서울"
    assert loaded.iloc[0]["담당자"] == "김철수"
    assert int(loaded.iloc[0]["매출"]) == 200000


def test_merged_header_base_supports_compound_names() -> None:
    assert merged_header_base("실행예산_2") == "실행예산"
    assert merged_header_base("실행예산_이월예산") == "실행예산"
    assert merged_header_base("계획예산") == "계획예산"
